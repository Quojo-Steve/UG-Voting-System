from datetime import datetime, date, time, timedelta
import smtplib
from email.message import EmailMessage
from fastapi import Depends, FastAPI, HTTPException, Header, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from io import BytesIO

from .config import get_settings
from .database import Base, engine, get_db
from .models import *
from .schemas import *
from .security import *

settings = get_settings()
app = FastAPI(title=settings.app_name)
app.add_middleware(
    CORSMiddleware,
    allow_origins=[o.strip() for o in settings.cors_origins.split(',')],
    allow_credentials=True,
    allow_methods=['*'],
    allow_headers=['*'],
)


def iso(dt):
    return dt.isoformat() + ('Z' if dt and dt.tzinfo is None else '') if dt else None


def parse_date(s: str) -> date:
    return date.fromisoformat(s)


def parse_time(s: str) -> time:
    return time.fromisoformat(s)


def mask_email(email: str) -> str:
    parts = email.split('@')
    if len(parts) != 2:
        return 'your registered email'
    user, domain = parts
    if len(user) <= 2:
        return f'{user[:1]}*@{domain}'
    return f'{user[0]}***{user[-1]}@{domain}'


def now_in_window(e: Election) -> bool:
    now = datetime.utcnow()
    start = datetime.combine(e.start_date, e.start_time)
    end = datetime.combine(e.end_date, e.end_time)
    return start <= now <= end


def election_counts(db: Session, eid: str):
    voters = db.scalar(select(func.count(Voter.id)).where(Voter.election_id == eid)) or 0
    votes = db.scalar(select(func.count(Ballot.id)).where(Ballot.election_id == eid)) or 0
    return voters, votes


def position_out(p: Position) -> PositionResponse:
    return PositionResponse(id=p.id, name=p.name, description=p.description, order=p.order)


def election_out(db: Session, e: Election) -> ElectionResponse:
    voters, ballots = election_counts(db, e.id)
    return ElectionResponse(
        id=e.id, name=e.name, description=e.description,
        startDate=e.start_date.isoformat(), startTime=e.start_time.strftime('%H:%M'),
        endDate=e.end_date.isoformat(), endTime=e.end_time.strftime('%H:%M'),
        status=e.status.value, positions=[position_out(p) for p in e.positions],
        totalRegisteredVoters=voters, totalVotesCast=ballots,
        createdAt=iso(e.created_at), publishedAt=iso(e.published_at),
    )


def candidate_votes(db: Session, cid: str) -> int:
    return db.scalar(select(func.count(Vote.id)).where(Vote.candidate_id == cid)) or 0


def candidate_out(db: Session, c: CandidateApplication) -> CandidateResponse:
    pos = db.get(Position, c.position_id)
    return CandidateResponse(
        id=c.id, electionId=c.election_id, positionId=c.position_id,
        positionName=pos.name if pos else '', fullName=c.full_name, email=c.email,
        studentId=c.student_id, hallOfResidence=c.hall_of_residence,
        department=c.department, level=c.level, manifesto=c.manifesto,
        runningMate=c.running_mate, avatarUrl=c.avatar_url, status=c.status.value,
        appliedAt=iso(c.applied_at), reviewNotes=c.review_notes,
        votesCount=candidate_votes(db, c.id),
    )


def voter_out(v: Voter) -> VoterResponse:
    return VoterResponse(
        id=v.id, voterId=v.voter_id, name=v.name, email=v.email, electionId=v.election_id,
        hall=v.hall, department=v.department, status=v.status.value, hasVoted=v.has_voted,
        votedAt=iso(v.voted_at), importedAt=iso(v.imported_at)
    )


def user_out(u: User) -> UserResponse:
    return UserResponse(
        id=u.id, email=u.email, name=u.name, role=u.role.value, studentId=u.student_id,
        fullName=u.name, hallOfResidence=u.hall_of_residence, department=u.department, level=u.level,
    )


def add_activity(db: Session, title: str, desc: str, typ: ActivityType, election_id: str | None = None, actor: User | None = None):
    db.add(ActivityLog(id=generate_id('act'), title=title, description=desc, type=typ, election_id=election_id, actor_user_id=actor.id if actor else None))


def send_otp_email(to_email: str, code: str):
    if not settings.smtp_enabled or not settings.smtp_host:
        print(f'[DEV OTP] {to_email}: {code}')
        return
    msg = EmailMessage()
    msg['Subject'] = 'UG Voting Verification OTP'
    msg['From'] = settings.smtp_from_email
    msg['To'] = to_email
    msg.set_content(f'Your University of Ghana voting OTP is {code}. It expires in {settings.otp_expire_minutes} minutes.')
    with smtplib.SMTP(settings.smtp_host, settings.smtp_port) as smtp:
        if settings.smtp_use_tls:
            smtp.starttls()
        if settings.smtp_username:
            smtp.login(settings.smtp_username, settings.smtp_password)
        smtp.send_message(msg)


def get_current_user(db: Session = Depends(get_db), authorization: str | None = Header(default=None)) -> User:
    if not authorization or not authorization.startswith('Bearer '):
        raise HTTPException(status_code=401, detail='Missing authentication token')
    try:
        payload = decode_token(authorization.split(' ', 1)[1])
    except Exception:
        raise HTTPException(status_code=401, detail='Invalid or expired token')
    user = db.get(User, payload.get('sub'))
    if not user or not user.is_active:
        raise HTTPException(status_code=401, detail='Invalid user')
    return user


def require_commissioner(user: User = Depends(get_current_user)) -> User:
    if user.role != Role.COMMISSIONER:
        raise HTTPException(status_code=403, detail='Commissioner role required')
    return user


def require_candidate(user: User = Depends(get_current_user)) -> User:
    if user.role != Role.CANDIDATE:
        raise HTTPException(status_code=403, detail='Candidate role required')
    return user


@app.on_event('startup')
def startup():
    Base.metadata.create_all(bind=engine)
    db = next(get_db())
    try:
        if not db.scalar(select(User).where(User.role == Role.COMMISSIONER)):
            db.add(User(id='comm-1', email='commissioner@ug.edu.gh', password_hash=hash_password('ecpassword2026'), role=Role.COMMISSIONER, name='University Electoral Commissioner'))
            db.add(ActivityLog(id=generate_id('act'), title='System Initialized', description='Default commissioner account created.', type=ActivityType.SYSTEM))
            db.commit()
    finally:
        db.close()


@app.get('/api/health')
def health():
    return {'status': 'ok'}


@app.post('/api/auth/commissioner/login', response_model=LoginResponse)
def login_commissioner(req: LoginRequest, db: Session = Depends(get_db)):
    u = db.scalar(select(User).where(User.email == req.email.lower()))
    if not u or u.role != Role.COMMISSIONER or not verify_password(req.password, u.password_hash):
        raise HTTPException(status_code=401, detail='Invalid email or password')
    return LoginResponse(token=create_access_token(u.id, u.role.value, u.email), user=user_out(u))


@app.post('/api/auth/candidate/login', response_model=LoginResponse)
def login_candidate(req: LoginRequest, db: Session = Depends(get_db)):
    u = db.scalar(select(User).where(User.email == req.email.lower()))
    if not u or u.role != Role.CANDIDATE or not verify_password(req.password, u.password_hash):
        raise HTTPException(status_code=401, detail='Invalid email or password')
    return LoginResponse(token=create_access_token(u.id, u.role.value, u.email), user=user_out(u))


@app.post('/api/auth/candidate/register', response_model=LoginResponse)
def register_candidate(req: CandidateRegisterRequest, db: Session = Depends(get_db)):
    if len(req.password) < 6:
        raise HTTPException(status_code=400, detail='Password must be at least 6 characters')
    u = User(id=generate_id('cand-user'), email=req.email.lower(), password_hash=hash_password(req.password), role=Role.CANDIDATE, name=req.fullName, student_id=req.studentId, hall_of_residence=req.hallOfResidence, department=req.department, level=req.level)
    db.add(u)
    try:
        add_activity(db, 'Candidate Registered', f'{req.fullName} registered a candidate account.', ActivityType.CANDIDATE)
        db.commit()
    except IntegrityError:
        db.rollback(); raise HTTPException(status_code=409, detail='Email or student ID already registered')
    return LoginResponse(token=create_access_token(u.id, u.role.value, u.email), user=user_out(u))


@app.get('/api/elections', response_model=list[ElectionResponse])
def list_elections(db: Session = Depends(get_db)):
    return [election_out(db, e) for e in db.scalars(select(Election).order_by(Election.created_at.desc())).all()]


@app.get('/api/elections/{eid}', response_model=ElectionResponse)
def get_election(eid: str, db: Session = Depends(get_db)):
    e = db.get(Election, eid)
    if not e: raise HTTPException(404, 'Election not found')
    return election_out(db, e)


@app.post('/api/elections', response_model=ElectionResponse)
def create_election(req: ElectionCreate, db: Session = Depends(get_db), user: User = Depends(require_commissioner)):
    try:
        sd, ed = parse_date(req.startDate), parse_date(req.endDate); st, et = parse_time(req.startTime), parse_time(req.endTime)
    except ValueError:
        raise HTTPException(400, 'Invalid date or time')
    if datetime.combine(ed, et) <= datetime.combine(sd, st):
        raise HTTPException(400, 'End date/time must be after start date/time')
    eid = generate_id('ele')
    e = Election(id=eid, name=req.name, description=req.description, start_date=sd, start_time=st, end_date=ed, end_time=et, status=ElectionStatus.DRAFT)
    db.add(e); db.flush()
    for i, name in enumerate(req.positions, 1):
        if name.strip(): db.add(Position(id=generate_id('pos'), election_id=eid, name=name.strip(), order=i))
    add_activity(db, 'Election Created', f'{req.name} created in DRAFT mode.', ActivityType.ELECTION, eid, user)
    db.commit(); db.refresh(e)
    return election_out(db, e)


@app.patch('/api/elections/{eid}/status', response_model=ElectionResponse)
def update_status(eid: str, req: StatusUpdate, db: Session = Depends(get_db), user: User = Depends(require_commissioner)):
    e = db.get(Election, eid)
    if not e: raise HTTPException(404, 'Election not found')
    try: new = ElectionStatus(req.status)
    except ValueError: raise HTTPException(400, 'Invalid status')
    allowed = {ElectionStatus.DRAFT:[ElectionStatus.REGISTRATION], ElectionStatus.REGISTRATION:[ElectionStatus.LIVE, ElectionStatus.READY], ElectionStatus.READY:[ElectionStatus.LIVE], ElectionStatus.LIVE:[ElectionStatus.CLOSED], ElectionStatus.CLOSED:[ElectionStatus.RESULTS_PUBLISHED]}
    if new != e.status and new not in allowed.get(e.status, []):
        raise HTTPException(409, 'Invalid election status transition')
    if new == ElectionStatus.LIVE:
        if db.scalar(select(func.count(Voter.id)).where(Voter.election_id == eid)) == 0: raise HTTPException(400, 'Cannot launch election without voters')
        if db.scalar(select(func.count(CandidateApplication.id)).where(CandidateApplication.election_id == eid, CandidateApplication.status == CandidateStatus.APPROVED)) == 0: raise HTTPException(400, 'Cannot launch election without approved candidates')
    e.status = new
    if new == ElectionStatus.RESULTS_PUBLISHED: e.published_at = datetime.utcnow()
    add_activity(db, f'Election Status Changed: {new.value}', f'{e.name} is now {new.value}', ActivityType.RESULT if new == ElectionStatus.RESULTS_PUBLISHED else ActivityType.ELECTION, eid, user)
    db.commit(); db.refresh(e)
    return election_out(db, e)


@app.get('/api/activity-logs', response_model=list[ActivityResponse])
def activity_logs(db: Session = Depends(get_db), user: User = Depends(require_commissioner)):
    rows = db.scalars(select(ActivityLog).order_by(ActivityLog.timestamp.desc()).limit(50)).all()
    return [ActivityResponse(id=a.id, title=a.title, description=a.description, timestamp=iso(a.timestamp), type=a.type.value) for a in rows]


@app.get('/api/candidates', response_model=list[CandidateResponse])
def all_candidates(db: Session = Depends(get_db), user: User = Depends(require_commissioner)):
    return [candidate_out(db, c) for c in db.scalars(select(CandidateApplication).order_by(CandidateApplication.applied_at.desc())).all()]


@app.get('/api/elections/{eid}/candidates', response_model=list[CandidateResponse])
def election_candidates(eid: str, db: Session = Depends(get_db)):
    if not db.get(Election, eid): raise HTTPException(404, 'Election not found')
    return [candidate_out(db, c) for c in db.scalars(select(CandidateApplication).where(CandidateApplication.election_id == eid).order_by(CandidateApplication.applied_at.desc())).all()]


@app.get('/api/candidates/me')
def candidates_me(email: str | None = None, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if user.role == Role.COMMISSIONER and email:
        return [candidate_out(db, c) for c in db.scalars(select(CandidateApplication).where(CandidateApplication.email == email.lower())).all()]
    if user.role != Role.CANDIDATE: raise HTTPException(403, 'Candidate role required')
    apps = db.scalars(select(CandidateApplication).where(CandidateApplication.user_id == user.id).order_by(CandidateApplication.applied_at.desc())).all()
    return candidate_out(db, apps[0]) if apps else None


@app.post('/api/elections/{eid}/candidates/apply', response_model=CandidateResponse)
def apply_candidate(eid: str, req: CandidateApply, db: Session = Depends(get_db), user: User = Depends(require_candidate)):
    e = db.get(Election, eid)
    if not e: raise HTTPException(404, 'Election not found')
    if e.status != ElectionStatus.REGISTRATION: raise HTTPException(409, 'Candidate registration is not open')
    pos = db.get(Position, req.positionId)
    if not pos or pos.election_id != eid: raise HTTPException(404, 'Position not found for election')
    c = CandidateApplication(id=generate_id('cand'), user_id=user.id, election_id=eid, position_id=req.positionId, full_name=req.fullName, email=req.email.lower(), student_id=req.studentId, hall_of_residence=req.hallOfResidence, department=req.department, level=req.level, manifesto=req.manifesto, running_mate=req.runningMate, avatar_url=req.avatarUrl, status=CandidateStatus.PENDING)
    db.add(c); add_activity(db, 'Candidate Application Submitted', f'{req.fullName} applied for {pos.name}', ActivityType.CANDIDATE, eid, user)
    try: db.commit()
    except IntegrityError: db.rollback(); raise HTTPException(409, 'Duplicate application for this election')
    return candidate_out(db, c)


def review_candidate_impl(cid: str, status_value: str, reason: str | None, db: Session, user: User):
    c = db.get(CandidateApplication, cid)
    if not c: raise HTTPException(404, 'Candidate not found')
    try: st = CandidateStatus(status_value)
    except ValueError: raise HTTPException(400, 'Invalid candidate status')
    c.status = st; c.review_notes = reason; c.reviewed_at = datetime.utcnow()
    add_activity(db, f'Candidate {st.value.title()}', f'{c.full_name} for application {c.id}', ActivityType.CANDIDATE, c.election_id, user)
    db.commit(); db.refresh(c)
    return candidate_out(db, c)


@app.patch('/api/candidates/{cid}/review', response_model=CandidateResponse)
def review_candidate(cid: str, req: CandidateReview, db: Session = Depends(get_db), user: User = Depends(require_commissioner)):
    return review_candidate_impl(cid, req.status, req.reason, db, user)


@app.patch('/api/candidates/{cid}/approve', response_model=CandidateResponse)
def approve_candidate(cid: str, db: Session = Depends(get_db), user: User = Depends(require_commissioner)):
    return review_candidate_impl(cid, 'APPROVED', None, db, user)


@app.patch('/api/candidates/{cid}/reject', response_model=CandidateResponse)
def reject_candidate(cid: str, req: RejectRequest, db: Session = Depends(get_db), user: User = Depends(require_commissioner)):
    return review_candidate_impl(cid, 'REJECTED', req.reason, db, user)


@app.get('/api/voters', response_model=list[VoterResponse])
def list_voters(db: Session = Depends(get_db), user: User = Depends(require_commissioner)):
    return [voter_out(v) for v in db.scalars(select(Voter).order_by(Voter.imported_at.desc())).all()]


@app.get('/api/elections/{eid}/voters', response_model=list[VoterResponse])
def election_voters(eid: str, db: Session = Depends(get_db), user: User = Depends(require_commissioner)):
    if not db.get(Election, eid): raise HTTPException(404, 'Election not found')
    return [voter_out(v) for v in db.scalars(select(Voter).where(Voter.election_id == eid).order_by(Voter.imported_at.desc())).all()]


@app.post('/api/elections/{eid}/voters/import', response_model=ImportResponse)
def import_voters(eid: str, req: VoterImportRequest, db: Session = Depends(get_db), user: User = Depends(require_commissioner)):
    e = db.get(Election, eid)
    if not e: raise HTTPException(404, 'Election not found')
    added = 0
    for item in req.voters:
        exists = db.scalar(select(Voter).where(Voter.election_id == eid, func.upper(Voter.voter_id) == item.voterId.strip().upper()))
        if exists: continue
        db.add(Voter(id=generate_id('voter'), election_id=eid, voter_id=item.voterId.strip(), name=item.name.strip(), email=item.email.lower(), hall=item.hall, department=item.department)); added += 1
    add_activity(db, 'Voter Register Uploaded', f'Imported {added} voter records into {e.name}', ActivityType.VOTER, eid, user)
    db.commit()
    return ImportResponse(count=added, importedCount=added, message=f'Successfully imported {added} student voters to register.')


def find_voter(db: Session, eid: str, voter_id: str) -> Voter | None:
    return db.scalar(select(Voter).where(Voter.election_id == eid, func.upper(Voter.voter_id) == voter_id.strip().upper()))


@app.post('/api/elections/{eid}/voters/import-file', response_model=ImportResponse)
def import_voters_file(eid: str, file: UploadFile = File(...), db: Session = Depends(get_db), user: User = Depends(require_commissioner)):
    """Optional Excel/CSV upload endpoint. The current frontend sends JSON, but this supports project-required openpyxl imports."""
    e = db.get(Election, eid)
    if not e: raise HTTPException(404, 'Election not found')
    content = file.file.read(); rows = []
    if file.filename.lower().endswith('.xlsx'):
        wb = load_workbook(BytesIO(content), read_only=True); ws = wb.active
        data = list(ws.iter_rows(values_only=True))
        headers = [str(x).strip().lower() if x is not None else '' for x in data[0]] if data else []
        for raw in data[1:]:
            row = {headers[i]: raw[i] for i in range(min(len(headers), len(raw)))}
            rows.append(VoterImportItem(voterId=str(row.get('voterid') or row.get('voter_id') or row.get('studentid') or ''), name=str(row.get('name') or ''), email=str(row.get('email') or ''), hall=str(row.get('hall') or '') or None, department=str(row.get('department') or '') or None))
    else:
        text = content.decode('utf-8-sig')
        lines = [ln for ln in text.splitlines() if ln.strip()]
        headers = [h.strip().lower() for h in lines[0].split(',')] if lines else []
        for ln in lines[1:]:
            vals = [v.strip() for v in ln.split(',')]
            row = {headers[i]: vals[i] for i in range(min(len(headers), len(vals)))}
            rows.append(VoterImportItem(voterId=row.get('voterid') or row.get('voter_id') or row.get('studentid') or '', name=row.get('name') or '', email=row.get('email') or '', hall=row.get('hall'), department=row.get('department')))
    return import_voters(eid, VoterImportRequest(voters=rows), db, user)


@app.post('/api/voter/request-otp', response_model=RequestOtpResponse)
def request_otp(req: RequestOtp, db: Session = Depends(get_db)):
    if not req.electionId: raise HTTPException(400, 'Election is required')
    e = db.get(Election, req.electionId)
    if not e: raise HTTPException(404, 'Election not found')
    v = find_voter(db, e.id, req.voterId)
    if not v: raise HTTPException(404, 'Voter ID was not found in the verified student register.')
    if v.has_voted: raise HTTPException(409, 'You have already voted in this election.')
    if v.status != VoterStatus.ELIGIBLE: raise HTTPException(403, 'Voter is not eligible')
    if e.status != ElectionStatus.LIVE: raise HTTPException(409, 'Voting is not currently live for this election.')
    code = generate_otp()
    db.add(OTPChallenge(id=generate_id('otp'), election_id=e.id, voter_db_id=v.id, otp_hash=hash_secret(code), expires_at=datetime.utcnow()+timedelta(minutes=settings.otp_expire_minutes)))
    db.commit(); send_otp_email(v.email, code)
    return RequestOtpResponse(success=True, maskedEmail=mask_email(v.email), electionId=e.id, electionName=e.name, voterName=v.name, message=f'Verification OTP sent to {mask_email(v.email)}')


@app.post('/api/voter/verify-otp', response_model=VoterSessionResponse)
def verify_otp(req: VerifyOtp, db: Session = Depends(get_db)):
    e = db.get(Election, req.electionId)
    if not e: raise HTTPException(404, 'Election not found')
    v = find_voter(db, e.id, req.voterId)
    if not v: raise HTTPException(404, 'Voter not found')
    rec = db.scalar(select(OTPChallenge).where(OTPChallenge.election_id == e.id, OTPChallenge.voter_db_id == v.id, OTPChallenge.used_at.is_(None)).order_by(OTPChallenge.created_at.desc()))
    if not rec: raise HTTPException(400, 'No active OTP request found')
    if datetime.utcnow() > rec.expires_at: raise HTTPException(410, 'Verification code has expired')
    if rec.otp_hash != hash_secret(req.code.strip()):
        rec.attempt_count += 1; db.commit(); raise HTTPException(401, 'Invalid OTP')
    if v.has_voted: raise HTTPException(409, 'You have already voted')
    if e.status != ElectionStatus.LIVE: raise HTTPException(409, 'Voting is not currently live')
    rec.used_at = datetime.utcnow(); token = generate_token(); expires = datetime.utcnow()+timedelta(minutes=settings.voting_session_expire_minutes)
    db.add(VotingSession(id=generate_id('vs'), token_hash=hash_secret(token), election_id=e.id, voter_db_id=v.id, expires_at=expires)); db.commit()
    return VoterSessionResponse(voterId=v.voter_id, name=v.name, voterName=v.name, maskedEmail=mask_email(v.email), electionId=e.id, electionName=e.name, verifiedOtpToken=token, token=token, expiresAt=iso(expires), hasVoted=v.has_voted)


@app.get('/api/elections/{eid}/ballot', response_model=BallotResponse)
def get_ballot(eid: str, db: Session = Depends(get_db)):
    e = db.get(Election, eid)
    if not e: raise HTTPException(404, 'Election not found')
    if e.status != ElectionStatus.LIVE: raise HTTPException(409, 'Voting is not currently live for this election')
    cands = db.scalars(select(CandidateApplication).where(CandidateApplication.election_id == eid, CandidateApplication.status == CandidateStatus.APPROVED)).all()
    return BallotResponse(election=election_out(db, e), positions=[position_out(p) for p in e.positions], candidates=[candidate_out(db, c) for c in cands])


@app.post('/api/elections/{eid}/ballot/cast', response_model=CastBallotResponse)
def cast_ballot(eid: str, req: CastBallotRequest, db: Session = Depends(get_db)):
    if eid != req.electionId: raise HTTPException(400, 'Election ID mismatch')
    e = db.get(Election, eid)
    if not e: raise HTTPException(404, 'Election not found')
    v = find_voter(db, eid, req.voterId)
    if not v: raise HTTPException(404, 'Voter not found')
    if not req.token: raise HTTPException(401, 'Missing voting token')
    sess = db.scalar(select(VotingSession).where(VotingSession.token_hash == hash_secret(req.token), VotingSession.election_id == eid, VotingSession.voter_db_id == v.id))
    if not sess or sess.used_at: raise HTTPException(401, 'Invalid voting token')
    if datetime.utcnow() > sess.expires_at: raise HTTPException(401, 'Voting token expired')
    if e.status != ElectionStatus.LIVE: raise HTTPException(409, 'Voting is not currently live')
    if v.has_voted: raise HTTPException(409, 'You have already voted')
    positions = db.scalars(select(Position).where(Position.election_id == eid)).all()
    pos_ids = {p.id for p in positions}; incoming = [x.positionId for x in req.votes]
    if set(incoming) != pos_ids or len(incoming) != len(set(incoming)): raise HTTPException(400, 'Incomplete or invalid ballot')
    ballot = Ballot(id=generate_id('ballot'), election_id=eid, voter_db_id=v.id, receipt_number=generate_receipt())
    db.add(ballot); db.flush()
    for item in req.votes:
        cand = db.get(CandidateApplication, item.candidateId)
        if not cand or cand.election_id != eid or cand.position_id != item.positionId or cand.status != CandidateStatus.APPROVED:
            raise HTTPException(400, 'Invalid candidate selection')
        db.add(Vote(id=generate_id('vote'), ballot_id=ballot.id, election_id=eid, position_id=item.positionId, candidate_id=cand.id))
    v.has_voted = True; v.status = VoterStatus.VOTED; v.voted_at = datetime.utcnow(); sess.used_at = datetime.utcnow()
    try: db.commit()
    except IntegrityError: db.rollback(); raise HTTPException(409, 'You have already voted in this election')
    return CastBallotResponse(success=True, receiptNumber=ballot.receipt_number, timestamp=iso(ballot.submitted_at))


def results_for(db: Session, e: Election) -> ElectionResults:
    registered, cast = election_counts(db, e.id)
    pos_results = []
    for p in e.positions:
        candidates = db.scalars(select(CandidateApplication).where(CandidateApplication.position_id == p.id, CandidateApplication.status == CandidateStatus.APPROVED)).all()
        rows = []
        for c in candidates:
            votes = candidate_votes(db, c.id); rows.append((c, votes))
        total = sum(v for _, v in rows); rows.sort(key=lambda x: x[1], reverse=True)
        cand_results = []
        for idx, (c, votes) in enumerate(rows, 1):
            cand_results.append(CandidateResult(candidateId=c.id, candidateName=c.full_name, positionId=p.id, positionName=p.name, avatarUrl=c.avatar_url, runningMate=c.running_mate, votes=votes, percentage=round((votes/total)*100, 1) if total else 0, rank=idx, isWinner=idx == 1 and votes > 0))
        pos_results.append(PositionResult(positionId=p.id, positionName=p.name, totalVotes=total, candidates=cand_results))
    turnout = round((cast/registered)*100, 1) if registered else 0
    return ElectionResults(electionId=e.id, electionName=e.name, status=e.status.value, isPublished=e.status == ElectionStatus.RESULTS_PUBLISHED, totalRegisteredVoters=registered, totalVotesCast=cast, turnoutPercentage=turnout, positions=pos_results, resultsByPosition=pos_results, publishedAt=iso(e.published_at))


@app.get('/api/elections/{eid}/results', response_model=ElectionResults)
def get_results(eid: str, db: Session = Depends(get_db), authorization: str | None = Header(default=None)):
    e = db.get(Election, eid)
    if not e: raise HTTPException(404, 'Election not found')
    is_comm = False
    if authorization and authorization.startswith('Bearer '):
        try:
            u = db.get(User, decode_token(authorization.split(' ',1)[1]).get('sub')); is_comm = bool(u and u.role == Role.COMMISSIONER)
        except Exception: pass
    if e.status != ElectionStatus.RESULTS_PUBLISHED and not is_comm:
        raise HTTPException(403, 'Results are not published')
    return results_for(db, e)


@app.post('/api/elections/{eid}/publish-results', response_model=ElectionResults)
def publish_results(eid: str, db: Session = Depends(get_db), user: User = Depends(require_commissioner)):
    e = db.get(Election, eid)
    if not e: raise HTTPException(404, 'Election not found')
    if e.status not in [ElectionStatus.CLOSED, ElectionStatus.RESULTS_PUBLISHED]: raise HTTPException(409, 'Election must be closed before publishing')
    e.status = ElectionStatus.RESULTS_PUBLISHED; e.published_at = e.published_at or datetime.utcnow()
    add_activity(db, 'Results Published', f'{e.name} results published.', ActivityType.RESULT, eid, user)
    db.commit(); return results_for(db, e)


@app.get('/api/candidates/me/results')
def candidate_result_summary(db: Session = Depends(get_db), user: User = Depends(require_candidate), email: str | None = None):
    elections = db.scalars(select(Election)).all()
    published = [results_for(db, e) for e in elections if e.status == ElectionStatus.RESULTS_PUBLISHED]
    unpublished = [{'electionId': e.id, 'electionName': e.name} for e in elections if e.status != ElectionStatus.RESULTS_PUBLISHED]
    return {'publishedResults': published, 'unpublishedElections': unpublished}
